#!/usr/bin/env python3
"""Extra spot checks for G1 docs provenance and G4 equality."""
from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DS = ROOT / "CS661_Dataset"
DASH = ROOT / "dashboard"
OUT = DASH / "_recon_g1_g4_extra_out.txt"


def parse_arr(path: Path, name: str):
    text = path.read_text(encoding="utf-8")
    m = re.search(rf"const {name}\s*=\s*", text)
    i = m.end()
    while text[i].isspace():
        i += 1
    depth = 0
    start = i
    in_str = False
    esc = False
    quote = ""
    for j in range(i, len(text)):
        ch = text[j]
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == quote:
                in_str = False
            continue
        if ch in "\"'":
            in_str = True
            quote = ch
            continue
        if ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                return json.loads(text[start : j + 1])
    raise ValueError(name)


def nearly(a, b, atol=0.05):
    return abs(a - b) <= atol


def main():
    buf = []

    master = list(csv.DictReader((DS / "master_dataset.csv").open(encoding="utf-8")))
    wb = list(csv.DictReader((DS / "raw_worldbank.csv").open(encoding="utf-8")))
    oa = list(csv.DictReader((DS / "raw_openalex_docs.csv").open(encoding="utf-8")))
    wbp = list(csv.DictReader((DS / "raw_wb_publications.csv").open(encoding="utf-8")))

    def idx(rows):
        return {(r["Country_Code"], int(r["Year"])): r for r in rows}

    mi, wi, oi, pi = idx(master), idx(wb), idx(oa), idx(wbp)

    buf.append("=== G1 docs/GDP/GERD provenance ===")
    for code in ["USA", "CHN", "IND", "GBR", "DEU"]:
        for y in [2015, 2020, 2022, 2024]:
            m, w, o, p = mi.get((code, y)), wi.get((code, y)), oi.get((code, y)), pi.get((code, y))
            buf.append(
                f"{code} {y}: master_docs={m.get('Total_Docs') if m else None} "
                f"OA={o.get('Total_Docs_OA') if o else None} "
                f"WB_pubs={p.get('WB_Pubs') if p else None}"
            )
            buf.append(
                f"         master_GDP={m.get('GDP_Per_Capita_PPP') if m else None} "
                f"rawWB_GDP={w.get('GDP_Per_Capita_PPP') if w else None} "
                f"master_GERD={m.get('GERD_Percent_GDP') if m else None} "
                f"rawWB_GERD={w.get('GERD_Percent_GDP') if w else None}"
            )

    q = defaultdict(set)
    for r in master:
        if r["Q1_Percent"]:
            q[r["Country_Code"]].add(r["Q1_Percent"])
    buf.append(
        f"countries with >1 distinct Q1_Percent: {sum(1 for c,s in q.items() if len(s)>1)} of {len(q)}"
    )
    buf.append(f"USA Q1 set: {q.get('USA')}")
    sjr = [
        r
        for r in csv.DictReader((DS / "sjr_country_metrics.csv").open(encoding="utf-8"))
        if r["Country_Code"] == "USA"
    ]
    buf.append(f"sjr USA: {sjr}")

    # imputed GERD for 2024
    try:
        import openpyxl

        wb_x = openpyxl.load_workbook(DS / "master_dataset_imputed.xlsx", read_only=True, data_only=True)
        ws = wb_x.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        ci = {h: i for i, h in enumerate(header)}
        buf.append("=== imputed xlsx GERD for checklist 2024 ===")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[ci["Country_Code"]] in {"USA", "CHN", "IND", "GBR", "DEU"} and row[ci["Year"]] == 2024:
                buf.append(
                    f"{row[ci['Country_Code']]} 2024 imputed GERD={row[ci['GERD_Percent_GDP']]} "
                    f"docs={row[ci['Total_Docs']]} H={row[ci['H_Index']]}"
                )
        wb_x.close()
    except Exception as e:
        buf.append(f"imputed xlsx error: {e}")

    # G4
    buf.append("\n=== G4 FE vs collaboration_premium ===")
    viz4 = parse_arr(DASH / "viz4_data.js", "VIZ4_DATA")
    prem = list(csv.DictReader((DS / "collaboration_premium.csv").open(encoding="utf-8")))
    by_name = defaultdict(list)
    for r in prem:
        by_name[r["Country_Name"]].append(r)

    fe_names = {r["name"] for r in viz4}
    prem_names = set(by_name)
    buf.append(f"FE n={len(viz4)} premium countries={len(prem_names)}")
    buf.append(f"overlap={sorted(fe_names & prem_names)}")
    buf.append(f"FE-only count={len(fe_names - prem_names)}")
    buf.append(f"premium-only={sorted(prem_names - fe_names)}")

    for name in sorted(fe_names & prem_names):
        fe = next(r for r in viz4 if r["name"] == name)
        rows = by_name[name]
        for label, sel in [
            ("mean", None),
            ("2022", next((r for r in rows if r["Year"] == "2022"), None)),
            ("2020", next((r for r in rows if r["Year"] == "2020"), None)),
            ("2015", next((r for r in rows if r["Year"] == "2015"), None)),
            ("latest", max(rows, key=lambda r: int(r["Year"]))),
            ("2010", next((r for r in rows if r["Year"] == "2010"), None)),
        ]:
            if label == "mean":
                dom = sum(float(r["Domestic_Avg_Citations"]) for r in rows) / len(rows)
                intl = sum(float(r["International_Avg_Citations"]) for r in rows) / len(rows)
                gain = sum(float(r["Citation_Gain"]) for r in rows) / len(rows)
            else:
                if not sel:
                    continue
                dom = float(sel["Domestic_Avg_Citations"])
                intl = float(sel["International_Avg_Citations"])
                gain = float(sel["Citation_Gain"])
            match = (
                nearly(fe["domestic"], round(dom, 1))
                and nearly(fe["international"], round(intl, 1))
                and nearly(fe["gain"], round(gain, 1))
            )
            match2 = nearly(fe["domestic"], dom, 0.2) and nearly(fe["international"], intl, 0.2)
            if match or match2 or label in ("2022", "mean", "latest"):
                buf.append(
                    f"{name:20s} {label:7s} FE={fe['domestic']}/{fe['international']}/{fe['gain']} "
                    f"SRC={dom:.3f}/{intl:.3f}/{gain:.3f} round1={match} near={match2}"
                )

    vals = []
    for r in viz4:
        for k in ["domestic", "international", "gain"]:
            vals.append(r[k])
    buf.append(f"all *10 integer? {all(abs(v*10 - round(v*10)) < 1e-9 for v in vals)}")
    gain_ok = sum(1 for r in viz4 if abs(r["gain"] - (r["international"] - r["domestic"])) < 0.05)
    buf.append(f"gain==intl-dom {gain_ok}/{len(viz4)}")

    # Compare FE USA to premium US 2022 exactly
    us_rows = [r for r in prem if r["Country_Code"] == "US"]
    buf.append("USA all premium years:")
    for r in sorted(us_rows, key=lambda x: x["Year"]):
        buf.append(
            f"  {r['Year']}: dom={r['Domestic_Avg_Citations']} intl={r['International_Avg_Citations']} "
            f"gain={r['Citation_Gain']}"
        )

    # G5 edge threshold: why IITK-IITI missing from year slices
    buf.append("\n=== G5 top-300 edge filter check ===")
    edges = list(
        csv.DictReader(
            (ROOT / "CS661 Project/data/processed/collaboration_edges_full.csv").open(encoding="utf-8")
        )
    )
    # cumulative all-years weights per pair
    pair_w = defaultdict(int)
    for e in edges:
        a, b = sorted([e["inst_a"], e["inst_b"]])
        pair_w[(a, b)] += int(e["weight"])
    top = sorted(pair_w.items(), key=lambda x: -x[1])[:300]
    min_top = top[-1][1] if top else None
    buf.append(f"top300 cumulative min weight={min_top} n_pairs={len(pair_w)}")
    iitk = "129be347-f8df-4fd2-b553-37ff0fd31af1"
    iiti = "92004806-cba2-4966-9ebf-9f31bc5a9597"
    key = tuple(sorted([iitk, iiti]))
    rank = next((i + 1 for i, (k, w) in enumerate(sorted(pair_w.items(), key=lambda x: -x[1])) if k == key), None)
    buf.append(f"IITK-IITI cumulative weight={pair_w[key]} rank={rank}")

    # year-slice top300 by single-year weight
    for year in ["2023", "2024"]:
        yw = defaultdict(int)
        for e in edges:
            if e["year"] == year:
                a, b = sorted([e["inst_a"], e["inst_b"]])
                yw[(a, b)] += int(e["weight"])
        top300 = sorted(yw.items(), key=lambda x: -x[1])[:300]
        thresh = top300[-1][1] if top300 else None
        w = yw.get(key, 0)
        in_top = any(k == key for k, _ in top300)
        buf.append(f"{year}: IITK-IITI weight={w} top300_min={thresh} in_top300={in_top} n_edges={len(yw)}")

    text = "\n".join(buf)
    OUT.write_text(text, encoding="utf-8")
    print(text)
    print(f"\nWrote {OUT}")


if __name__ == "__main__":
    main()
